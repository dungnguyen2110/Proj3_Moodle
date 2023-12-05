import unittest
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl


class TestAssignSystemRoleBonus(unittest.TestCase):
    def test_AssignSystemRole(self):
        self.workbook = openpyxl.load_workbook(filename="./testcase.xlsx")
        self.sheet = self.workbook["EditCourse_FullNameBonus"]

        for row in range(2, self.sheet.max_row + 1):
            print("In testcase number " + str(row - 1))
            get_value = lambda column: self.sheet.cell(row=row, column=column).value

            try:
                options = webdriver.ChromeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Chrome(options=options)
            except:
                options = webdriver.EdgeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Edge(options=options)

            self.driver.set_window_size(1460, 1080)
            self.driver.get(get_value(1))

            # Find username and password
            username = self.driver.find_element(By.ID, get_value(2))
            password = self.driver.find_element(By.ID, get_value(3))

            # Enter credentials
            username.send_keys(get_value(4))
            password.send_keys(get_value(5))

            # Submit form
            submit = self.driver.find_element(By.ID, get_value(6))
            submit.click()
            time.sleep(3)

             # Click on Course
            self.driver.find_element(By.CSS_SELECTOR, get_value(7)).click()

             #  Click on Settings
            self.driver.find_element(By.LINK_TEXT, get_value(8)).click()

             # Click on Full Name
            fullname = self.driver.find_element(By.ID, get_value(9))

            # Get the value from the Excel sheet
            cell_value = self.sheet.cell(row=row, column=11).value
            
            # Check if the value is None and set it to an empty string if True
            if cell_value is None:
                cell_value = ""
            fullname.clear()
        
            # Enter course name
            fullname.send_keys(cell_value)
            
            time.sleep(3)

            # Save and display
            self.driver.find_element(By.ID, get_value(10)).click()
            print('Testcase number ' + str(row - 1) + ' of Edit Course General Settings- Course Full Name has finished.')
            print("==========================================")
            self.driver.quit()

    def tearDown(self):
        print(f"\n{self._testMethodName} finished\n")


if __name__ == "__main__":
    unittest.main()
