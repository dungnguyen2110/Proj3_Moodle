import unittest
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl


class TestAssignSystemRole(unittest.TestCase):
    def test_AssignSystemRole(self):
        self.workbook = openpyxl.load_workbook(filename="./testcase.xlsx")
        self.sheet = self.workbook["EditCourse_FullName"]

        for row in range(2, self.sheet.max_row + 1):
            print("In testcase number " + str(row - 1))

            try:
                options = webdriver.ChromeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Chrome(options=options)
            except:
                options = webdriver.EdgeOptions()
                options.add_argument("--log-level=3")
                self.driver = webdriver.Edge(options=options)

            self.driver.set_window_size(1460, 1080)
            self.driver.get("https://school.moodledemo.net/login/index.php")

            # Find username and password
            username = self.driver.find_element(By.ID, "username")
            password = self.driver.find_element(By.ID, "password")

            # Enter credentials
            username.send_keys(self.sheet.cell(row=row, column=1).value)
            password.send_keys(self.sheet.cell(row=row, column=2).value)

            # Submit form
            submit = self.driver.find_element(By.ID, "loginbtn")
            submit.click()
            time.sleep(3)

            # Click on Course
            self.driver.find_element(By.CSS_SELECTOR, ".multiline").click()

            #  Click on Settings
            self.driver.find_element(By.LINK_TEXT, "Settings").click()
            

            # Click on Full Name
            
            fullname = self.driver.find_element(By.ID, "id_fullname")
            
            # Get the value from the Excel sheet
            cell_value = self.sheet.cell(row=row, column=3).value
            
            # Check if the value is None and set it to an empty string if True
            if cell_value is None:
                cell_value = ""
            fullname.clear()
        
            # Enter course name
            fullname.send_keys(cell_value)
            
            time.sleep(3)
            
            # Save and display
            self.driver.find_element(By.ID, "id_saveanddisplay").click()
            print('Testcase number ' + str(row - 1) + ' of Edit Course General Settings- Course Full Name has finished.')
            print("==========================================")
            self.driver.quit()

    def tearDown(self):
        print(f"\n{self._testMethodName} finished\n")


if __name__ == "__main__":
    unittest.main()
